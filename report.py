# -*- coding: utf-8 -*-

import openpyxl as xl
import paramiko
import sys
import json

def _set_cell(ws, row, column, value, error=False):
    ws.cell(row=row, column=column).value = value.decode("ISO-8859-1")
    ws.cell(row=row, column=column).alignment = xl.styles.Alignment(wrapText=True)
    if error:
        redFill = xl.styles.PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        ws.cell(row=row, column=column).fill = redFill

def _connect_host(host, username, password, wait_time, ip_gateway=None, port=22):

    if ip_gateway:
        vm = paramiko.SSHClient()
        vm.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        vm.connect(ip_gateway, username=username, password=password, auth_timeout=wait_time)
        vmtransport = vm.get_transport()
        vmchannel = vmtransport.open_channel("direct-tcpip", (host, port), (ip_gateway, port))
        jhost = paramiko.SSHClient()
        jhost.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jhost.connect(host, username=username, password=password, sock=vmchannel, auth_timeout=wait_time)

    else:
        jhost = paramiko.SSHClient()
        jhost.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jhost.connect(host, username=username, password=password, auth_timeout=wait_time)

    stdin, hostname, stderr1 = jhost.exec_command('uname -n 2>/dev/null')
    stdin, oslevel, stderr2 = jhost.exec_command('oslevel -s 2>/dev/null || uname -r 2>/dev/null')
    hostname = hostname.read()
    oslevel = oslevel.read()
    if not hostname:
        hostname = '###### {}'.format(stderr1.read())
    if not oslevel:
        oslevel = '###### {}'.format(stderr2.read())
    jhost.close()
    if ip_gateway:
        vm.close()

    values = {}
    values['hostname'] = hostname
    values['oslevel'] = oslevel
    return values


if __name__ == "__main__":
    info = json.load(open('access.json'))

    filename = info['file'] # .xlsx file
    gateway = info['gateway'] # Gateway used to reach final host
    username = info['username'] # Username for both Gateway and final host
    password = info['password'] # Password for both Gateway and final host
    wait_time = info['wait_time'] # Waiting time to try connection with Gateway/host
    col_ip = info['col_ip'] # number of column on .xlsx where IPs are found
    col_start = info['col_start'] # in case host not found, which columns should be  written with -
    col_end = info['col_end'] # in case host not found, which columns should be  written with -
    port = info['port'] # port used for both connections
    wb = xl.load_workbook(filename)
    ws = wb.active

    for col in ws.iter_cols(min_row=2, min_col=col_ip, max_col=col_ip):
        for cell in col:
            if not cell.value:
                print 'FINISHED PROCESSING AT ROW {}\n'.format(cell.row),
                sys.stdout.flush()
                sys.exit(0)
            print 'Processing row {}: '.format(cell.row),
            sys.stdout.flush()
            if not ws.cell(row=cell.row, column=col_start).value:
                print 'CONNECTING to {}, '.format(str(cell.value)),
                sys.stdout.flush()
                try:
                    values = _connect_host(host=str(cell.value),\
                            username = username, password = password,\
                            wait_time = wait_time,\
                            ip_gateway = gateway, port = port)
                    _set_cell(ws, cell.row, 3, values['hostname'])
                    _set_cell(ws, cell.row, 4, values['oslevel'])
                    print "DONE\n".format(cell.row),
                    sys.stdout.flush()
                except Exception as e:
                    for i in range(col_start,col_end+1):
                        _set_cell(ws, cell.row, i, str(e), True)
                    print "CAN'T CONNECT - {}\n".format(str(e)),
                    sys.stdout.flush()
                wb.save(filename)
            else:
                print "DONE\n".format(cell.row),
sys.stdout.flush()
